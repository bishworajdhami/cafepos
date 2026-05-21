#!/usr/bin/env python3
"""Script to read and update the Gantt chart Excel file"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
import os

# Path to the Excel file
file_path = r"c:\Users\bishw\source\repos\cafeSystem\cafeSystem\Source\cafe\src\Report\cafe_gantt_chart.xlsx"

# Load the workbook
wb = load_workbook(file_path)
ws = wb.active

print("Current Gantt Chart Content:")
print("=" * 100)

# Display current content
for row in ws.iter_rows(values_only=False):
    for cell in row:
        if cell.value:
            print(f"Cell {cell.coordinate}: {cell.value}", end=" | ")
    print()

print("\n" + "=" * 100)
print("Current data (showing values only):")
print("=" * 100)

# Show current data
for i, row in enumerate(ws.iter_rows(max_row=ws.max_row, values_only=True), 1):
    print(f"Row {i}: {row}")

print("\n" + "=" * 100)
print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
