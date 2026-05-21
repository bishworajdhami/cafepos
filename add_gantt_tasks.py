#!/usr/bin/env python3
"""Script to add more tasks to the Gantt chart"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Path to the Excel file
file_path = r"c:\Users\bishw\source\repos\cafeSystem\cafeSystem\Source\cafe\src\Report\cafe_gantt_chart.xlsx"

# Load the workbook
wb = load_workbook(file_path)
ws = wb.active

# New tasks to add (Task Name, Module, Start Week, Duration in Weeks)
new_tasks = [
    ("Table Booking System", "Waiter", 7, 2),
    ("Waiter Management", "Waiter", 6, 3),
    ("Customer Feedback Module", "Manager/Admin", 10, 2),
    ("Inventory Management", "Manager/Admin", 7, 2),
    ("Security Audit & Enhancement", "Infrastructure", 8, 2),
    ("Performance Optimization", "Infrastructure", 11, 2),
    ("API Documentation", "Infrastructure", 11, 1),
    ("User Training & Onboarding", "Manager/Admin", 12, 1),
    ("Bug Fixes & Refinement", "Infrastructure", 13, 2),
    ("Post-Deployment Support", "Infrastructure", 14, 2),
    ("Mobile App Optimization", "Infrastructure", 9, 2),
    ("Real-time Notifications", "Kitchen/Chef", 8, 2),
]

# Get the current last row
current_last_row = ws.max_row

# Get the header styling from the first row
header_font = ws['A1'].font
header_fill = ws['A1'].fill

# Add new tasks
for idx, (task_name, module, start_week, duration) in enumerate(new_tasks, start=1):
    new_row = current_last_row + idx
    
    # Add data
    ws[f'A{new_row}'] = task_name
    ws[f'B{new_row}'] = module
    ws[f'C{new_row}'] = start_week
    ws[f'D{new_row}'] = duration
    
    # Apply formatting (copy from row 2)
    for col in ['A', 'B', 'C', 'D']:
        source_cell = ws[f'{col}2']
        target_cell = ws[f'{col}{new_row}']
        
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color
            )
        
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )
        
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )

# Save the updated workbook
wb.save(file_path)

print(f"✓ Successfully added {len(new_tasks)} new tasks to the Gantt chart!")
print(f"✓ Total tasks now: {ws.max_row - 1} (including header)")
print("\nNew tasks added:")
print("-" * 80)
for idx, (task_name, module, start_week, duration) in enumerate(new_tasks, start=1):
    end_week = start_week + duration - 1
    print(f"{idx:2d}. {task_name:40s} | {module:20s} | Week {start_week}-{end_week} ({duration}w)")

print("-" * 80)
print(f"\nFile saved: {file_path}")
